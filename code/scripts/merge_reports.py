"""
Script for merging CSV comparison reports into Excel files.

This script takes CSV comparison results and merges them into
consolidated Excel workbooks for easier analysis.
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from ..config.settings import load_config, get_config
from ..core.constants import (
    TIMESTAMP_FORMAT,
    EXCEL_HEADER_COLOR,
    EXCEL_SEPARATOR_COLOR,
    EXCEL_MAX_COLUMN_WIDTH
)
from ..utils.file_handler import FileHandler, CSVHandler, ExcelHandler
from ..utils.logger import get_logger, shutdown_logging, PerformanceLogger
from ..utils.validators import PathValidator


class MergeService:
    """
    Service for merging CSV comparison reports into Excel files.
    """
    
    def __init__(self, logger):
        """
        Initialize merge service.
        
        Args:
            logger: Logger instance
        """
        self.logger = logger
        self.config = get_config()
    
    def group_csv_files(self, csv_files: List[Path]) -> Dict[str, List[Path]]:
        """
        Group CSV files by common prefix.
        
        Args:
            csv_files: List of CSV file paths
            
        Returns:
            Dictionary mapping group names to file lists
        """
        file_groups = {}
        
        for file in csv_files:
            # Extract prefix (everything before first underscore)
            if '_' in file.stem:
                prefix = file.stem.split('_')[0]
            else:
                prefix = file.stem
            
            if prefix not in file_groups:
                file_groups[prefix] = []
            file_groups[prefix].append(file)
        
        return file_groups
    
    def merge_csv_group(
        self,
        files: List[Path],
        group_name: str,
        output_folder: Path
    ) -> Path:
        """
        Merge a group of CSV files into a single Excel workbook.
        
        Args:
            files: List of CSV files to merge
            group_name: Name for the merged file
            output_folder: Output folder
            
        Returns:
            Path to created Excel file
        """
        try:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{group_name}_merged"
            
            # Define styles
            header_fill = PatternFill(
                start_color=EXCEL_HEADER_COLOR,
                end_color=EXCEL_HEADER_COLOR,
                fill_type="solid"
            )
            separator_fill = PatternFill(
                start_color=EXCEL_SEPARATOR_COLOR,
                end_color=EXCEL_SEPARATOR_COLOR,
                fill_type="solid"
            )
            header_font = Font(color="FFFFFF", bold=True)
            
            current_row = 1
            header_added = False
            
            # Process each CSV file
            for csv_file in sorted(files):
                try:
                    # Read CSV
                    df = CSVHandler.read_csv_to_dataframe(csv_file)
                    
                    if df.empty:
                        self.logger.warning(f"Empty CSV file: {csv_file.name}")
                        continue
                    
                    # Add header if first file
                    if not header_added:
                        headers = ['Source_File'] + list(df.columns)
                        ws.append(headers)
                        
                        # Style header row
                        for cell in ws[current_row]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        current_row += 1
                        header_added = True
                    
                    # Add data rows
                    for _, row in df.iterrows():
                        data_row = [csv_file.name] + list(row)
                        ws.append(data_row)
                        current_row += 1
                    
                    # Add separator row
                    separator_row = [''] * (len(df.columns) + 1)
                    ws.append(separator_row)
                    
                    # Style separator
                    for cell in ws[current_row]:
                        cell.fill = separator_fill
                    
                    current_row += 1
                    
                    self.logger.debug(f"Added {len(df)} rows from {csv_file.name}")
                    
                except Exception as e:
                    self.logger.error(f"Failed to process {csv_file.name}: {e}")
                    continue
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, EXCEL_MAX_COLUMN_WIDTH)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            output_file = output_folder / f"{group_name}.xlsx"
            ExcelHandler.save_workbook(wb, output_file)
            
            self.logger.info(f"Merged {len(files)} files into {output_file.name}")
            return output_file
            
        except Exception as e:
            self.logger.error(f"Failed to merge group {group_name}: {e}")
            raise
    
    def merge_comparison_folder(
        self,
        comparison_folder: Path,
        output_folder: Optional[Path] = None
    ) -> dict:
        """
        Merge all CSV files in a comparison folder.
        
        Args:
            comparison_folder: Folder containing CSV comparison results
            output_folder: Output folder for merged files (auto-created if None)
            
        Returns:
            Dictionary with merge results
        """
        try:
            with PerformanceLogger(self.logger, f"Merge CSV files from {comparison_folder.name}"):
                # Validate input folder
                PathValidator.validate_directory_exists(
                    comparison_folder,
                    "Comparison folder"
                )
                
                # Get CSV files
                csv_files = FileHandler.list_files(comparison_folder, "*.csv")
                
                if not csv_files:
                    self.logger.warning(f"No CSV files found in {comparison_folder}")
                    return {
                        'input_folder': str(comparison_folder),
                        'output_folder': None,
                        'total_csv_files': 0,
                        'file_groups': 0,
                        'successful_merges': 0,
                        'failed_merges': 0,
                        'merged_files': []
                    }
                
                self.logger.info(f"Found {len(csv_files)} CSV files to merge")
                
                # Group files
                file_groups = self.group_csv_files(csv_files)
                self.logger.info(f"Grouped into {len(file_groups)} groups")
                
                # Create output folder
                if output_folder is None:
                    timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
                    output_folder = comparison_folder / f"merged_{timestamp}"
                
                FileHandler.ensure_directory(output_folder)
                
                # Merge each group
                merged_files = []
                successful = 0
                failed = 0
                
                for group_name, files in file_groups.items():
                    try:
                        output_file = self.merge_csv_group(
                            files=files,
                            group_name=group_name,
                            output_folder=output_folder
                        )
                        merged_files.append({
                            'group_name': group_name,
                            'file_count': len(files),
                            'output_file': str(output_file),
                            'success': True
                        })
                        successful += 1
                    except Exception as e:
                        self.logger.error(f"Failed to merge group {group_name}: {e}")
                        merged_files.append({
                            'group_name': group_name,
                            'file_count': len(files),
                            'output_file': None,
                            'success': False,
                            'error': str(e)
                        })
                        failed += 1
                
                # Generate summary
                results = {
                    'input_folder': str(comparison_folder),
                    'output_folder': str(output_folder),
                    'total_csv_files': len(csv_files),
                    'file_groups': len(file_groups),
                    'successful_merges': successful,
                    'failed_merges': failed,
                    'merged_files': merged_files
                }
                
                # Write summary file
                self._write_merge_summary(results, output_folder)
                
                self.logger.info(
                    f"Merge completed: {successful}/{len(file_groups)} groups successful"
                )
                
                return results
                
        except Exception as e:
            self.logger.error(f"Merge operation failed: {e}")
            raise
    
    def _write_merge_summary(self, results: dict, output_folder: Path) -> None:
        """Write merge summary to text file."""
        summary_file = output_folder / "merge_summary.txt"
        
        lines = [
            "CSV MERGE SUMMARY",
            "=" * 80,
            "",
            f"Input Folder: {results['input_folder']}",
            f"Output Folder: {results['output_folder']}",
            f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "STATISTICS:",
            f"  Total CSV files: {results['total_csv_files']}",
            f"  File groups: {results['file_groups']}",
            f"  Successful merges: {results['successful_merges']}",
            f"  Failed merges: {results['failed_merges']}",
            "",
            "MERGE RESULTS:"
        ]
        
        for merge_result in results['merged_files']:
            status = "✓" if merge_result['success'] else "✗"
            lines.append(
                f"  {status} {merge_result['group_name']}: {merge_result['file_count']} files"
            )
            if merge_result['success']:
                lines.append(f"    Output: {merge_result['output_file']}")
            else:
                lines.append(f"    Error: {merge_result.get('error', 'Unknown error')}")
        
        FileHandler.write_text_file(summary_file, "\n".join(lines))
        self.logger.info(f"Summary written to: {summary_file}")


class MergeRunner:
    """
    Orchestrates CSV merge operations.
    """
    
    def __init__(self, config_file: Optional[Path] = None):
        """
        Initialize merge runner.
        
        Args:
            config_file: Path to configuration file
        """
        # Load configuration
        self.config = load_config(config_file, validate_paths=False)
        
        # Setup logging
        log_file = (
            self.config.paths.logs /
            f"merge_{datetime.now().strftime(TIMESTAMP_FORMAT)}.log"
        )
        self.logger = get_logger(
            name="MergeRunner",
            log_file=log_file,
            level=self.config.logging.level
        )
        
        # Initialize service
        self.merge_service = MergeService(self.logger)
        
        self.logger.info("Merge Runner initialized")
    
    def merge_folder(
        self,
        folder_name: str,
        base_folder: Optional[Path] = None
    ) -> dict:
        """
        Merge CSV files from a comparison folder.
        
        Args:
            folder_name: Name of comparison folder to merge
            base_folder: Base comparisons folder (uses config if None)
            
        Returns:
            Dictionary with merge results
        """
        try:
            # Use config path if base folder not provided
            if base_folder is None:
                base_folder = self.config.paths.output_comparisons
            
            # Build full path
            comparison_folder = base_folder / folder_name
            
            self.logger.info(f"Merging CSV files from: {comparison_folder}")
            
            # Perform merge
            results = self.merge_service.merge_comparison_folder(comparison_folder)
            
            return results
            
        except Exception as e:
            self.logger.error(f"Merge failed: {e}")
            raise
        finally:
            shutdown_logging()


def main():
    """Main entry point with CLI argument parsing."""
    parser = argparse.ArgumentParser(
        description="API Testing Framework - Merge CSV comparison reports",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Merge CSV files from a comparison folder
  python merge_reports.py pre_vs_post
  
  # Merge with custom base folder
  python merge_reports.py results --base-folder /path/to/comparisons
  
  # Use custom config file
  python merge_reports.py results --config /path/to/config.json
        """
    )
    
    parser.add_argument(
        'folder_name',
        type=str,
        help='Name of the comparison folder to merge'
    )
    
    parser.add_argument(
        '--base-folder',
        type=Path,
        help='Base comparisons folder (default: from config)'
    )
    
    parser.add_argument(
        '--config',
        type=Path,
        help='Path to configuration file (default: config.json)'
    )
    
    args = parser.parse_args()
    
    try:
        # Initialize and run merge
        runner = MergeRunner(config_file=args.config)
        results = runner.merge_folder(
            folder_name=args.folder_name,
            base_folder=args.base_folder
        )
        
        # Print summary
        print("\n" + "=" * 80)
        print("MERGE SUMMARY")
        print("=" * 80)
        print(f"Input Folder: {results['input_folder']}")
        print(f"Output Folder: {results['output_folder']}")
        print(f"\nStatistics:")
        print(f"  Total CSV files: {results['total_csv_files']}")
        print(f"  File groups: {results['file_groups']}")
        print(f"  Successful merges: {results['successful_merges']}")
        print(f"  Failed merges: {results['failed_merges']}")
        
        if results['successful_merges'] > 0:
            print(f"\nMerged Files:")
            for merge_result in results['merged_files']:
                if merge_result['success']:
                    print(f"  ✓ {merge_result['group_name']}: {merge_result['output_file']}")
        
        if results['failed_merges'] > 0:
            print(f"\nFailed Merges:")
            for merge_result in results['merged_files']:
                if not merge_result['success']:
                    print(f"  ✗ {merge_result['group_name']}: {merge_result.get('error', 'Unknown')}")
        
        print("=" * 80)
        
        return 0 if results['failed_merges'] == 0 else 1
        
    except KeyboardInterrupt:
        print("\n\nMerge interrupted by user")
        return 130
    except Exception as e:
        print(f"\n\nERROR: Merge failed: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())