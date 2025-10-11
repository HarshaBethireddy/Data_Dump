"""
Centralized file I/O operations with robust error handling.

This module provides utilities for reading, writing, and managing files
with proper encoding detection, error handling, and path validation.
"""

import json
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import aiofiles
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as WorkbookType

from apitesting.core.constants import (
    DEFAULT_ENCODING,
    FALLBACK_ENCODINGS,
    FileExtension
)
from apitesting.core.exceptions import (
    FileOperationError,
    JSONProcessingError,
    ExcelProcessingError
)


class FileHandler:
    """Handles file I/O operations with error handling and encoding detection."""
    
    @staticmethod
    def ensure_directory(path: Union[str, Path]) -> Path:
        """
        Ensure a directory exists, creating it if necessary.
        
        Args:
            path: Directory path
            
        Returns:
            Path object for the directory
            
        Raises:
            FileOperationError: If directory creation fails
        """
        try:
            path = Path(path)
            path.mkdir(parents=True, exist_ok=True)
            return path
        except Exception as e:
            raise FileOperationError(
                f"Failed to create directory: {path}",
                file_path=str(path),
                operation="create_directory",
                original_error=e
            )
    
    @staticmethod
    def read_text_file(
        file_path: Union[str, Path],
        encoding: Optional[str] = None
    ) -> str:
        """
        Read text file with automatic encoding detection.
        
        Args:
            file_path: Path to the file
            encoding: Specific encoding to use (auto-detect if None)
            
        Returns:
            File contents as string
            
        Raises:
            FileOperationError: If file cannot be read
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileOperationError(
                f"File not found: {file_path}",
                file_path=str(file_path),
                operation="read"
            )
        
        encodings = [encoding] if encoding else [DEFAULT_ENCODING, *FALLBACK_ENCODINGS]
        
        for enc in encodings:
            try:
                with open(file_path, 'r', encoding=enc) as f:
                    return f.read()
            except UnicodeDecodeError:
                continue
            except Exception as e:
                raise FileOperationError(
                    f"Failed to read file: {file_path}",
                    file_path=str(file_path),
                    operation="read",
                    original_error=e
                )
        
        raise FileOperationError(
            f"Could not read file with any supported encoding: {file_path}",
            file_path=str(file_path),
            operation="read",
            details={"attempted_encodings": list(encodings)}
        )
    
    @staticmethod
    def write_text_file(
        file_path: Union[str, Path],
        content: str,
        encoding: str = DEFAULT_ENCODING,
        create_dirs: bool = True
    ) -> None:
        """
        Write text content to file.
        
        Args:
            file_path: Path to the file
            content: Content to write
            encoding: File encoding
            create_dirs: Whether to create parent directories
            
        Raises:
            FileOperationError: If file cannot be written
        """
        file_path = Path(file_path)
        
        try:
            if create_dirs:
                FileHandler.ensure_directory(file_path.parent)
            
            with open(file_path, 'w', encoding=encoding) as f:
                f.write(content)
        except Exception as e:
            raise FileOperationError(
                f"Failed to write file: {file_path}",
                file_path=str(file_path),
                operation="write",
                original_error=e
            )
    
    @staticmethod
    async def read_text_file_async(
        file_path: Union[str, Path],
        encoding: str = DEFAULT_ENCODING
    ) -> str:
        """
        Asynchronously read text file.
        
        Args:
            file_path: Path to the file
            encoding: File encoding
            
        Returns:
            File contents as string
            
        Raises:
            FileOperationError: If file cannot be read
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileOperationError(
                f"File not found: {file_path}",
                file_path=str(file_path),
                operation="read_async"
            )
        
        try:
            async with aiofiles.open(file_path, 'r', encoding=encoding) as f:
                return await f.read()
        except Exception as e:
            raise FileOperationError(
                f"Failed to read file asynchronously: {file_path}",
                file_path=str(file_path),
                operation="read_async",
                original_error=e
            )
    
    @staticmethod
    async def write_text_file_async(
        file_path: Union[str, Path],
        content: str,
        encoding: str = DEFAULT_ENCODING,
        create_dirs: bool = True
    ) -> None:
        """
        Asynchronously write text content to file.
        
        Args:
            file_path: Path to the file
            content: Content to write
            encoding: File encoding
            create_dirs: Whether to create parent directories
            
        Raises:
            FileOperationError: If file cannot be written
        """
        file_path = Path(file_path)
        
        try:
            if create_dirs:
                FileHandler.ensure_directory(file_path.parent)
            
            async with aiofiles.open(file_path, 'w', encoding=encoding) as f:
                await f.write(content)
        except Exception as e:
            raise FileOperationError(
                f"Failed to write file asynchronously: {file_path}",
                file_path=str(file_path),
                operation="write_async",
                original_error=e
            )
    
    @staticmethod
    def copy_file(
        source: Union[str, Path],
        destination: Union[str, Path],
        create_dirs: bool = True
    ) -> None:
        """
        Copy file from source to destination.
        
        Args:
            source: Source file path
            destination: Destination file path
            create_dirs: Whether to create parent directories
            
        Raises:
            FileOperationError: If file cannot be copied
        """
        source = Path(source)
        destination = Path(destination)
        
        if not source.exists():
            raise FileOperationError(
                f"Source file not found: {source}",
                file_path=str(source),
                operation="copy"
            )
        
        try:
            if create_dirs:
                FileHandler.ensure_directory(destination.parent)
            
            shutil.copy2(source, destination)
        except Exception as e:
            raise FileOperationError(
                f"Failed to copy file from {source} to {destination}",
                file_path=str(source),
                operation="copy",
                original_error=e
            )
    
    @staticmethod
    def delete_file(file_path: Union[str, Path], missing_ok: bool = True) -> None:
        """
        Delete a file.
        
        Args:
            file_path: Path to the file
            missing_ok: If True, don't raise error if file doesn't exist
            
        Raises:
            FileOperationError: If file cannot be deleted
        """
        file_path = Path(file_path)
        
        try:
            file_path.unlink(missing_ok=missing_ok)
        except Exception as e:
            raise FileOperationError(
                f"Failed to delete file: {file_path}",
                file_path=str(file_path),
                operation="delete",
                original_error=e
            )
    
    @staticmethod
    def list_files(
        directory: Union[str, Path],
        pattern: str = "*",
        recursive: bool = False
    ) -> List[Path]:
        """
        List files in a directory matching a pattern.
        
        Args:
            directory: Directory to search
            pattern: Glob pattern to match
            recursive: Whether to search recursively
            
        Returns:
            List of matching file paths
            
        Raises:
            FileOperationError: If directory cannot be read
        """
        directory = Path(directory)
        
        if not directory.exists():
            raise FileOperationError(
                f"Directory not found: {directory}",
                file_path=str(directory),
                operation="list_files"
            )
        
        try:
            if recursive:
                return sorted(directory.rglob(pattern))
            else:
                return sorted(directory.glob(pattern))
        except Exception as e:
            raise FileOperationError(
                f"Failed to list files in directory: {directory}",
                file_path=str(directory),
                operation="list_files",
                original_error=e
            )


class JSONHandler:
    """Handles JSON file operations with validation and error handling."""
    
    @staticmethod
    def read_json(
        file_path: Union[str, Path],
        encoding: str = DEFAULT_ENCODING
    ) -> Union[Dict, List]:
        """
        Read and parse JSON file.
        
        Args:
            file_path: Path to JSON file
            encoding: File encoding
            
        Returns:
            Parsed JSON data
            
        Raises:
            JSONProcessingError: If JSON cannot be read or parsed
        """
        file_path = Path(file_path)
        
        try:
            content = FileHandler.read_text_file(file_path, encoding)
            return json.loads(content)
        except json.JSONDecodeError as e:
            raise JSONProcessingError(
                f"Invalid JSON in file: {file_path}",
                file_path=str(file_path),
                details={"error": str(e), "line": e.lineno, "column": e.colno},
                original_error=e
            )
        except Exception as e:
            raise JSONProcessingError(
                f"Failed to read JSON file: {file_path}",
                file_path=str(file_path),
                original_error=e
            )
    
    @staticmethod
    def write_json(
        file_path: Union[str, Path],
        data: Union[Dict, List],
        indent: int = 4,
        encoding: str = DEFAULT_ENCODING,
        create_dirs: bool = True
    ) -> None:
        """
        Write data to JSON file.
        
        Args:
            file_path: Path to JSON file
            data: Data to write
            indent: JSON indentation
            encoding: File encoding
            create_dirs: Whether to create parent directories
            
        Raises:
            JSONProcessingError: If JSON cannot be written
        """
        file_path = Path(file_path)
        
        try:
            json_content = json.dumps(data, indent=indent, ensure_ascii=False)
            FileHandler.write_text_file(file_path, json_content, encoding, create_dirs)
        except Exception as e:
            raise JSONProcessingError(
                f"Failed to write JSON file: {file_path}",
                file_path=str(file_path),
                original_error=e
            )
    
    @staticmethod
    async def read_json_async(
        file_path: Union[str, Path],
        encoding: str = DEFAULT_ENCODING
    ) -> Union[Dict, List]:
        """
        Asynchronously read and parse JSON file.
        
        Args:
            file_path: Path to JSON file
            encoding: File encoding
            
        Returns:
            Parsed JSON data
            
        Raises:
            JSONProcessingError: If JSON cannot be read or parsed
        """
        file_path = Path(file_path)
        
        try:
            content = await FileHandler.read_text_file_async(file_path, encoding)
            return json.loads(content)
        except json.JSONDecodeError as e:
            raise JSONProcessingError(
                f"Invalid JSON in file: {file_path}",
                file_path=str(file_path),
                details={"error": str(e), "line": e.lineno, "column": e.colno},
                original_error=e
            )
        except Exception as e:
            raise JSONProcessingError(
                f"Failed to read JSON file asynchronously: {file_path}",
                file_path=str(file_path),
                original_error=e
            )
    
    @staticmethod
    async def write_json_async(
        file_path: Union[str, Path],
        data: Union[Dict, List],
        indent: int = 4,
        encoding: str = DEFAULT_ENCODING,
        create_dirs: bool = True
    ) -> None:
        """
        Asynchronously write data to JSON file.
        
        Args:
            file_path: Path to JSON file
            data: Data to write
            indent: JSON indentation
            encoding: File encoding
            create_dirs: Whether to create parent directories
            
        Raises:
            JSONProcessingError: If JSON cannot be written
        """
        file_path = Path(file_path)
        
        try:
            json_content = json.dumps(data, indent=indent, ensure_ascii=False)
            await FileHandler.write_text_file_async(file_path, json_content, encoding, create_dirs)
        except Exception as e:
            raise JSONProcessingError(
                f"Failed to write JSON file asynchronously: {file_path}",
                file_path=str(file_path),
                original_error=e
            )


class ExcelHandler:
    """Handles Excel file operations with validation and error handling."""
    
    @staticmethod
    def load_workbook(file_path: Union[str, Path]) -> WorkbookType:
        """
        Load Excel workbook.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Loaded workbook
            
        Raises:
            ExcelProcessingError: If workbook cannot be loaded
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise ExcelProcessingError(
                f"Excel file not found: {file_path}",
                file_path=str(file_path)
            )
        
        try:
            return load_workbook(file_path)
        except Exception as e:
            raise ExcelProcessingError(
                f"Failed to load Excel workbook: {file_path}",
                file_path=str(file_path),
                original_error=e
            )
    
    @staticmethod
    def save_workbook(
        workbook: WorkbookType,
        file_path: Union[str, Path],
        create_dirs: bool = True
    ) -> None:
        """
        Save Excel workbook.
        
        Args:
            workbook: Workbook to save
            file_path: Path to save file
            create_dirs: Whether to create parent directories
            
        Raises:
            ExcelProcessingError: If workbook cannot be saved
        """
        file_path = Path(file_path)
        
        try:
            if create_dirs:
                FileHandler.ensure_directory(file_path.parent)
            
            workbook.save(file_path)
        except Exception as e:
            raise ExcelProcessingError(
                f"Failed to save Excel workbook: {file_path}",
                file_path=str(file_path),
                original_error=e
            )
    
    @staticmethod
    def read_excel_to_dataframe(
        file_path: Union[str, Path],
        sheet_name: Union[str, int] = 0
    ) -> pd.DataFrame:
        """
        Read Excel file into pandas DataFrame.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or index
            
        Returns:
            DataFrame with Excel data
            
        Raises:
            ExcelProcessingError: If Excel cannot be read
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise ExcelProcessingError(
                f"Excel file not found: {file_path}",
                file_path=str(file_path)
            )
        
        try:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception as e:
            raise ExcelProcessingError(
                f"Failed to read Excel file to DataFrame: {file_path}",
                file_path=str(file_path),
                sheet_name=str(sheet_name),
                original_error=e
            )
    
    @staticmethod
    def write_dataframe_to_excel(
        df: pd.DataFrame,
        file_path: Union[str, Path],
        sheet_name: str = "Sheet1",
        create_dirs: bool = True
    ) -> None:
        """
        Write pandas DataFrame to Excel file.
        
        Args:
            df: DataFrame to write
            file_path: Path to Excel file
            sheet_name: Sheet name
            create_dirs: Whether to create parent directories
            
        Raises:
            ExcelProcessingError: If DataFrame cannot be written
        """
        file_path = Path(file_path)
        
        try:
            if create_dirs:
                FileHandler.ensure_directory(file_path.parent)
            
            df.to_excel(file_path, sheet_name=sheet_name, index=False)
        except Exception as e:
            raise ExcelProcessingError(
                f"Failed to write DataFrame to Excel: {file_path}",
                file_path=str(file_path),
                sheet_name=sheet_name,
                original_error=e
            )


class CSVHandler:
    """Handles CSV file operations with validation and error handling."""
    
    @staticmethod
    def read_csv_to_dataframe(
        file_path: Union[str, Path],
        encoding: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Read CSV file into pandas DataFrame with automatic encoding detection.
        
        Args:
            file_path: Path to CSV file
            encoding: Specific encoding (auto-detect if None)
            
        Returns:
            DataFrame with CSV data
            
        Raises:
            FileOperationError: If CSV cannot be read
        """
        file_path = Path(file_path)
        
        if not file_path.exists():
            raise FileOperationError(
                f"CSV file not found: {file_path}",
                file_path=str(file_path),
                operation="read_csv"
            )
        
        encodings = [encoding] if encoding else [DEFAULT_ENCODING, *FALLBACK_ENCODINGS]
        
        for enc in encodings:
            try:
                return pd.read_csv(file_path, encoding=enc)
            except UnicodeDecodeError:
                continue
            except Exception as e:
                raise FileOperationError(
                    f"Failed to read CSV file: {file_path}",
                    file_path=str(file_path),
                    operation="read_csv",
                    original_error=e
                )
        
        raise FileOperationError(
            f"Could not read CSV with any supported encoding: {file_path}",
            file_path=str(file_path),
            operation="read_csv",
            details={"attempted_encodings": list(encodings)}
        )
    
    @staticmethod
    def write_dataframe_to_csv(
        df: pd.DataFrame,
        file_path: Union[str, Path],
        encoding: str = DEFAULT_ENCODING,
        create_dirs: bool = True
    ) -> None:
        """
        Write pandas DataFrame to CSV file.
        
        Args:
            df: DataFrame to write
            file_path: Path to CSV file
            encoding: File encoding
            create_dirs: Whether to create parent directories
            
        Raises:
            FileOperationError: If DataFrame cannot be written
        """
        file_path = Path(file_path)
        
        try:
            if create_dirs:
                FileHandler.ensure_directory(file_path.parent)
            
            df.to_csv(file_path, encoding=encoding, index=False)
        except Exception as e:
            raise FileOperationError(
                f"Failed to write DataFrame to CSV: {file_path}",
                file_path=str(file_path),
                operation="write_csv",
                original_error=e
            )