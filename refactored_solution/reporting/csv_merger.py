"""
Enhanced CSV merger with better error handling and formatting.
"""
import os
import pandas as pd
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import logging

from ..utils.logger import framework_logger
from ..utils.run_manager import RunManager


class CSVMerger:
    """Enhanced CSV merger with improved formatting and error handling."""
    
    def __init__(self, output_folder: str = "MergedOut_File"):
        self.output_folder = output_folder
        self.logger = framework_logger.get_logger()
        self.run_manager = RunManager()
    
    def merge_csv_files(self, parent_folder: str, sub_folder: str) -> Dict[str, Any]:
        """
        Merge CSV files from a subfolder into Excel workbooks.
        
        Args:
            parent_folder: Parent folder containing subfolders with CSV files
            sub_folder: Specific subfolder containing CSV files to merge
            
        Returns:
            Dictionary containing merge results and statistics
        """
        try:
            self.logger.info(f"Starting CSV merge for folder: {sub_folder}")
            
            # Validate input folder
            input_folder = os.path.join(parent_folder, sub_folder)
            if not os.path.exists(input_folder):
                raise FileNotFoundError(f"Input folder not found: {input_folder}")
            
            # Get CSV files
            csv_files = [f for f in os.listdir(input_folder) if f.endswith('.csv')]
            if not csv_files:
                raise ValueError(f"No CSV files found in {input_folder}")
            
            self.logger.info(f"Found {len(csv_files)} CSV files to merge")
            
            # Group files by common prefix
            file_groups = self._group_files_by_prefix(csv_files)
            
            # Create timestamped output folder
            timestamp = self.run_manager.get_timestamp()
            output_subfolder = os.path.join(self.output_folder, timestamp)
            os.makedirs(output_subfolder, exist_ok=True)
            
            # Merge each group
            merge_results = []
            for common_name, files in file_groups.items():
                try:
                    output_file = self._merge_file_group(
                        input_folder, files, output_subfolder, common_name
                    )
                    merge_results.append({
                        'group_name': common_name,
                        'file_count': len(files),
                        'output_file': output_file,
                        'success': True
                    })
                    self.logger.info(f"Successfully merged {len(files)} files into {common_name}.xlsx")
                except Exception as e:
                    self.logger.error(f"Failed to merge group {common_name}: {e}")
                    merge_results.append({
                        'group_name': common_name,
                        'file_count': len(files),
                        'output_file': None,
                        'success': False,
                        'error': str(e)
                    })
            
            # Generate summary
            summary = {
                'input_folder': input_folder,
                'output_folder': output_subfolder,
                'total_csv_files': len(csv_files),
                'file_groups': len(file_groups),
                'successful_merges': sum(1 for r in merge_results if r['success']),
                'failed_merges': sum(1 for r in merge_results if not r['success']),
                'merge_results': merge_results,
                'timestamp': timestamp
            }
            
            self._generate_merge_summary(summary)
            
            self.logger.info(f"CSV merge completed: {summary['successful_merges']}/{len(file_groups)} groups successful")
            return summary
            
        except Exception as e:
            self.logger.error(f"CSV merge failed: {e}")
            raise
    
    def _group_files_by_prefix(self, csv_files: List[str]) -> Dict[str, List[str]]:
        """Group CSV files by their common prefix (before first underscore)."""
        file_groups = {}
        
        for file in csv_files:
            # Extract prefix (everything before first underscore)
            if '_' in file:
                prefix = file.split('_')[0]
            else:
                # If no underscore, use filename without extension
                prefix = os.path.splitext(file)[0]
            
            if prefix not in file_groups:
                file_groups[prefix] = []
            file_groups[prefix].append(file)
        
        return file_groups
    
    def _merge_file_group(self, input_folder: str, files: List[str], output_folder: str, group_name: str) -> str:
        """Merge a group of CSV files into a single Excel workbook."""
        try:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"{group_name}_merged"
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            separator_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            # Track if header has been added
            header_added = False
            current_row = 1
            
            for file_name in sorted(files):
                try:
                    # Read CSV file with multiple encoding attempts
                    df = self._read_csv_safely(os.path.join(input_folder, file_name))
                    
                    if df.empty:
                        self.logger.warning(f"Empty CSV file: {file_name}")
                        continue
                    
                    # Add header row if first file
                    if not header_added:
                        headers = ['Source_File'] + list(df.columns)
                        ws.append(headers)
                        
                        # Style header row
                        for cell in ws[current_row]:
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')
                        
                        current_row += 1
                        header_added = True
                    
                    # Add data rows
                    for _, row in df.iterrows():
                        data_row = [file_name] + list(row)
                        ws.append(data_row)
                        current_row += 1
                    
                    # Add separator row
                    separator_row = [''] * (len(df.columns) + 1)
                    ws.append(separator_row)
                    
                    # Style separator row
                    for cell in ws[current_row]:
                        cell.fill = separator_fill
                    
                    current_row += 1
                    
                    self.logger.debug(f"Added {len(df)} rows from {file_name}")
                    
                except Exception as e:
                    self.logger.error(f"Failed to process file {file_name}: {e}")
                    continue
            
            # Auto-adjust column widths
            self._auto_adjust_columns(ws)
            
            # Save workbook
            output_file = os.path.join(output_folder, f"{group_name}.xlsx")
            wb.save(output_file)
            
            return output_file
            
        except Exception as e:
            self.logger.error(f"Failed to merge file group {group_name}: {e}")
            raise
    
    def _read_csv_safely(self, file_path: str) -> pd.DataFrame:
        """Read CSV file with multiple encoding attempts."""
        encodings = ['utf-8', 'utf-8-sig', 'iso-8859-1', 'cp1252', 'latin1']
        
        for encoding in encodings:
            try:
                return pd.read_csv(file_path, encoding=encoding)
            except UnicodeDecodeError:
                continue
            except Exception as e:
                self.logger.error(f"Failed to read {file_path} with {encoding}: {e}")
                continue
        
        raise RuntimeError(f"Could not read CSV file {file_path} with any supported encoding")
    
    def _auto_adjust_columns(self, worksheet) -> None:
        """Auto-adjust column widths based on content."""
        try:
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set width with some padding, but cap at reasonable maximum
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.warning(f"Failed to auto-adjust columns: {e}")
    
    def _generate_merge_summary(self, summary: Dict[str, Any]) -> None:
        """Generate a summary report of the merge operation."""
        try:
            summary_file = os.path.join(summary['output_folder'], "merge_summary.txt")
            
            with open(summary_file, 'w', encoding='utf-8') as file:
                file.write("CSV MERGE SUMMARY REPORT\n")
                file.write("=" * 40 + "\n\n")
                
                file.write(f"Input Folder: {summary['input_folder']}\n")
                file.write(f"Output Folder: {summary['output_folder']}\n")
                file.write(f"Timestamp: {summary['timestamp']}\n\n")
                
                file.write("STATISTICS:\n")
                file.write(f"  Total CSV files: {summary['total_csv_files']}\n")
                file.write(f"  File groups: {summary['file_groups']}\n")
                file.write(f"  Successful merges: {summary['successful_merges']}\n")
                file.write(f"  Failed merges: {summary['failed_merges']}\n\n")
                
                file.write("MERGE RESULTS:\n")
                for result in summary['merge_results']:
                    status = "✓" if result['success'] else "✗"
                    file.write(f"  {status} {result['group_name']}: {result['file_count']} files\n")
                    if result['success']:
                        file.write(f"    Output: {result['output_file']}\n")
                    else:
                        file.write(f"    Error: {result.get('error', 'Unknown error')}\n")
            
            self.logger.info(f"Merge summary generated: {summary_file}")
            
        except Exception as e:
            self.logger.error(f"Failed to generate merge summary: {e}")


def merge_comparison_results(sub_folder: str, parent_folder: str = "CompareResult") -> Dict[str, Any]:
    """
    Merge CSV comparison results into Excel workbooks.
    
    Args:
        sub_folder: Subfolder containing CSV files to merge
        parent_folder: Parent folder containing the subfolder
        
    Returns:
        Dictionary containing merge results and statistics
    """
    try:
        merger = CSVMerger()
        return merger.merge_csv_files(parent_folder, sub_folder)
        
    except Exception as e:
        logger = framework_logger.get_logger()
        logger.error(f"Failed to merge comparison results: {e}")
        raise