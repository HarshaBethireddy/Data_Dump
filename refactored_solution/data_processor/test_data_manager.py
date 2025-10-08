"""
Unified test data management for both regular and prequal test data.
"""
import os
import json
from typing import List, Optional, Union
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from decimal import Decimal
import logging

from ..utils.logger import framework_logger


class TestDataManager:
    """Unified manager for test data processing."""
    
    def __init__(self, source_folder: str, excel_file: str, destination_folder: str = "json_files"):
        self.source_folder = source_folder
        self.excel_file = excel_file
        self.destination_folder = destination_folder
        self.logger = framework_logger.get_logger()
        
        # Determine if this is prequal data based on file name
        self.is_prequal = "prequal" in excel_file.lower()
    
    def validate_inputs(self) -> None:
        """Validate that required inputs exist."""
        if not os.path.exists(self.source_folder):
            raise FileNotFoundError(f"Source folder not found: {self.source_folder}")
        
        if not os.path.exists(self.excel_file):
            raise FileNotFoundError(f"Excel file not found: {self.excel_file}")
        
        # Check if source folder has JSON files
        json_files = [f for f in os.listdir(self.source_folder) if f.endswith('.json')]
        if not json_files:
            raise ValueError(f"No JSON files found in source folder: {self.source_folder}")
    
    def _load_workbook_safely(self) -> Workbook:
        """Load Excel workbook with error handling."""
        try:
            return load_workbook(self.excel_file)
        except Exception as e:
            raise RuntimeError(f"Failed to load Excel file {self.excel_file}: {e}")
    
    def _get_last_appid_value(self, sheet) -> Union[int, str]:
        """Get the last APPID value from Excel sheet."""
        try:
            last_row = sheet.max_row
            last_value = sheet.cell(row=last_row, column=1).value
            
            if last_value is None:
                raise ValueError("No data found in Excel file")
            
            if self.is_prequal:
                # For prequal, handle as string to preserve leading zeros
                return str(last_value)
            else:
                # For regular data, convert to int
                return int(last_value)
        except (ValueError, TypeError) as e:
            raise ValueError(f"Invalid APPID value in Excel file: {e}")
    
    def _increment_appid(self, last_value: Union[int, str], position: int) -> Union[int, str]:
        """Increment APPID value based on type."""
        if self.is_prequal:
            # Handle 20-digit numbers as strings to avoid overflow
            try:
                # Convert to Decimal for safe arithmetic
                base_value = Decimal(last_value.strip())
                new_value = base_value + Decimal(position + 1)
                # Format as 20-digit string with leading zeros
                return f"{new_value:020d}"
            except Exception as e:
                raise ValueError(f"Failed to increment prequal APPID: {e}")
        else:
            # Regular integer increment
            return int(last_value) + position + 1
    
    def update_excel_file(self) -> None:
        """Update Excel file with incremented APPID values."""
        try:
            workbook = self._load_workbook_safely()
            sheet = workbook.active
            
            # Get the last value
            last_value = self._get_last_appid_value(sheet)
            self.logger.info(f"Last APPID value in Excel: {last_value}")
            
            # Update Excel column with incremented values
            row = 2  # Start from A2
            position = 0
            
            while sheet.cell(row=row, column=1).value is not None:
                new_value = self._increment_appid(last_value, position)
                sheet.cell(row=row, column=1).value = new_value
                row += 1
                position += 1
            
            # Save the updated Excel file
            workbook.save(self.excel_file)
            self.logger.info(f"Excel file updated successfully with {position} records")
            
        except Exception as e:
            raise RuntimeError(f"Failed to update Excel file: {e}")
    
    def _read_json_file_safely(self, file_path: str) -> dict:
        """Read JSON file with error handling."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return json.load(file)
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in file {file_path}: {e}")
        except Exception as e:
            raise RuntimeError(f"Failed to read file {file_path}: {e}")
    
    def _replace_appid_in_json(self, data: dict, appid_value: Union[int, str]) -> dict:
        """Replace $APPID placeholder in JSON data."""
        try:
            # Convert to string for replacement
            json_str = json.dumps(data)
            json_str = json_str.replace("$APPID", str(appid_value))
            return json.loads(json_str)
        except Exception as e:
            raise RuntimeError(f"Failed to replace APPID in JSON: {e}")
    
    def process_json_files(self) -> List[str]:
        """Process all JSON files with updated APPID values."""
        try:
            # Ensure destination folder exists
            os.makedirs(self.destination_folder, exist_ok=True)
            
            # Load Excel file to get APPID values
            workbook = self._load_workbook_safely()
            sheet = workbook.active
            
            # Get sorted list of JSON files
            json_files = sorted([f for f in os.listdir(self.source_folder) if f.endswith('.json')])
            processed_files = []
            
            for idx, filename in enumerate(json_files):
                try:
                    # Get APPID value from Excel
                    appid_cell = sheet.cell(row=idx + 2, column=1).value
                    
                    if appid_cell is None:
                        self.logger.warning(f"No APPID found for row {idx + 2}. Skipping file {filename}")
                        continue
                    
                    # Read and process JSON file
                    source_path = os.path.join(self.source_folder, filename)
                    data = self._read_json_file_safely(source_path)
                    
                    # Replace APPID placeholder
                    updated_data = self._replace_appid_in_json(data, appid_cell)
                    
                    # Save updated JSON file
                    destination_path = os.path.join(self.destination_folder, filename)
                    with open(destination_path, 'w', encoding='utf-8') as file:
                        json.dump(updated_data, file, indent=4, ensure_ascii=False)
                    
                    processed_files.append(filename)
                    self.logger.info(f"Processed file: {filename} with APPID: {appid_cell}")
                    
                except Exception as e:
                    self.logger.error(f"Failed to process file {filename}: {e}")
                    continue
            
            self.logger.info(f"Successfully processed {len(processed_files)} JSON files")
            return processed_files
            
        except Exception as e:
            raise RuntimeError(f"Failed to process JSON files: {e}")
    
    def run_full_process(self) -> List[str]:
        """Run the complete test data processing workflow."""
        try:
            self.logger.info(f"Starting test data processing for {self.source_folder}")
            
            # Validate inputs
            self.validate_inputs()
            
            # Update Excel file
            self.update_excel_file()
            
            # Process JSON files
            processed_files = self.process_json_files()
            
            self.logger.info("Test data processing completed successfully")
            return processed_files
            
        except Exception as e:
            self.logger.error(f"Test data processing failed: {e}")
            raise


def create_regular_test_data_manager(destination_folder: str = "json_files") -> TestDataManager:
    """Create a test data manager for regular test data."""
    return TestDataManager(
        source_folder="FullSetRequest",
        excel_file="MasterTestdata.xlsx",
        destination_folder=destination_folder
    )


def create_prequal_test_data_manager(destination_folder: str = "json_files") -> TestDataManager:
    """Create a test data manager for prequal test data."""
    return TestDataManager(
        source_folder="Prequal Request",
        excel_file="PreQual_MasterTestdata.xlsx",
        destination_folder=destination_folder
    )