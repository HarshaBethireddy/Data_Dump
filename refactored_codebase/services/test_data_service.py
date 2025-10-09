"""
Test data preparation service.

Handles processing of JSON templates with Excel data to generate
test payloads for API testing.
"""

import json
from pathlib import Path
from typing import List, Optional, Union, Dict, Any
from decimal import Decimal
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import logging

from ..core.config import ConfigurationManager


class TestDataProcessor:
    """
    Base class for test data processing operations.
    
    Provides common functionality for reading Excel data and
    processing JSON templates.
    """
    
    def __init__(self, 
                 config_manager: ConfigurationManager,
                 logger: logging.Logger):
        """
        Initialize test data processor.
        
        Args:
            config_manager: Configuration manager instance
            logger: Logger instance
        """
        self.config_manager = config_manager
        self.logger = logger
        self.path_config = config_manager.path_config
    
    def read_excel_data(self, excel_file: Path, sheet_name: str = "Sheet1") -> List[Dict[str, Any]]:
        """
        Read data from Excel file.
        
        Args:
            excel_file: Path to Excel file
            sheet_name: Name of the sheet to read
            
        Returns:
            List of dictionaries containing row data
            
        Raises:
            FileNotFoundError: If Excel file doesn't exist
            RuntimeError: If reading fails
        """
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_file}")
        
        try:
            workbook = load_workbook(excel_file, data_only=True)
            
            if sheet_name not in workbook.sheetnames:
                available_sheets = ", ".join(workbook.sheetnames)
                raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
            
            worksheet = workbook[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value is not None:
                    headers.append(str(cell.value).strip())
                else:
                    headers.append(f"Column_{len(headers) + 1}")
            
            # Read data rows
            data = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_data = {}
                    for i, value in enumerate(row):
                        if i < len(headers):
                            # Handle large numbers safely
                            if isinstance(value, (int, float)) and len(str(value)) > 15:
                                row_data[headers[i]] = str(Decimal(str(value)))
                            else:
                                row_data[headers[i]] = value
                    data.append(row_data)
            
            self.logger.debug(f"Read {len(data)} rows from {excel_file}")
            return data
            
        except Exception as e:
            raise RuntimeError(f"Failed to read Excel file {excel_file}: {e}") from e
    
    def load_json_template(self, template_file: Path) -> Dict[str, Any]:
        """
        Load JSON template file.
        
        Args:
            template_file: Path to JSON template file
            
        Returns:
            Parsed JSON data
            
        Raises:
            FileNotFoundError: If template file doesn't exist
            RuntimeError: If JSON parsing fails
        """
        if not template_file.exists():
            raise FileNotFoundError(f"Template file not found: {template_file}")
        
        try:
            with open(template_file, 'r', encoding='utf-8') as file:
                return json.load(file)
        except json.JSONDecodeError as e:
            raise RuntimeError(f"Invalid JSON in template {template_file}: {e}") from e
        except Exception as e:
            raise RuntimeError(f"Failed to load template {template_file}: {e}") from e
    
    def substitute_placeholders(self, template: Dict[str, Any], data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Substitute placeholders in JSON template with actual data.
        
        Args:
            template: JSON template with placeholders
            data: Data to substitute
            
        Returns:
            JSON with substituted values
        """
        template_str = json.dumps(template)
        
        # Substitute placeholders
        for key, value in data.items():
            placeholder = f"${key}"
            if placeholder in template_str:
                # Handle different data types appropriately
                if isinstance(value, str):
                    template_str = template_str.replace(placeholder, value)
                else:
                    template_str = template_str.replace(f'"{placeholder}"', str(value))
        
        return json.loads(template_str)
    
    def save_processed_json(self, data: Dict[str, Any], output_file: Path) -> None:
        """
        Save processed JSON data to file.
        
        Args:
            data: JSON data to save
            output_file: Output file path
        """
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as file:
            json.dump(data, file, indent=2, ensure_ascii=False)


class RegularTestDataProcessor(TestDataProcessor):
    """Processor for regular (fullset) test data."""
    
    def process(self) -> List[str]:
        """
        Process regular test data.
        
        Returns:
            List of processed file names
        """
        self.logger.info("Processing regular test data")
        
        # Read Excel data
        excel_data = self.read_excel_data(self.path_config.master_testdata_file)
        
        # Get JSON templates
        template_files = list(self.path_config.fullset_request_dir.glob("*.json"))
        
        if not template_files:
            self.logger.warning(f"No JSON templates found in {self.path_config.fullset_request_dir}")
            return []
        
        processed_files = []
        
        for template_file in template_files:
            self.logger.debug(f"Processing template: {template_file.name}")
            template = self.load_json_template(template_file)
            
            for i, row_data in enumerate(excel_data, 1):
                # Generate unique APPID
                app_id = f"{i:06d}"
                row_data['APPID'] = app_id
                
                # Substitute placeholders
                processed_json = self.substitute_placeholders(template, row_data)
                
                # Save processed file
                output_file = (self.path_config.json_output_dir / 
                             f"{template_file.stem}_{app_id}.json")
                self.save_processed_json(processed_json, output_file)
                processed_files.append(output_file.name)
        
        self.logger.info(f"Processed {len(processed_files)} regular test files")
        return processed_files


class PrequalTestDataProcessor(TestDataProcessor):
    """Processor for prequal test data."""
    
    def process(self) -> List[str]:
        """
        Process prequal test data.
        
        Returns:
            List of processed file names
        """
        self.logger.info("Processing prequal test data")
        
        # Read Excel data
        excel_data = self.read_excel_data(self.path_config.prequal_testdata_file)
        
        # Get JSON templates
        template_files = list(self.path_config.prequal_request_dir.glob("*.json"))
        
        if not template_files:
            self.logger.warning(f"No JSON templates found in {self.path_config.prequal_request_dir}")
            return []
        
        processed_files = []
        
        for template_file in template_files:
            self.logger.debug(f"Processing template: {template_file.name}")
            template = self.load_json_template(template_file)
            
            for i, row_data in enumerate(excel_data, 1):
                # Generate unique APPID (20-digit for prequal)
                app_id = f"{i:020d}"
                row_data['APPID'] = app_id
                
                # Substitute placeholders
                processed_json = self.substitute_placeholders(template, row_data)
                
                # Save processed file
                output_file = (self.path_config.json_output_dir / 
                             f"{template_file.stem}_{app_id}.json")
                self.save_processed_json(processed_json, output_file)
                processed_files.append(output_file.name)
        
        self.logger.info(f"Processed {len(processed_files)} prequal test files")
        return processed_files


class TestDataService:
    """
    High-level service for test data preparation.
    
    Orchestrates different test data processors and provides
    a unified interface for test data preparation.
    """
    
    def __init__(self, 
                 config_manager: ConfigurationManager,
                 logger: logging.Logger):
        """
        Initialize test data service.
        
        Args:
            config_manager: Configuration manager instance
            logger: Logger instance
        """
        self.config_manager = config_manager
        self.logger = logger
        
        # Initialize processors
        self.regular_processor = RegularTestDataProcessor(config_manager, logger)
        self.prequal_processor = PrequalTestDataProcessor(config_manager, logger)
    
    def prepare_test_data(self, data_type: str = "both") -> List[str]:
        """
        Prepare test data based on specified type.
        
        Args:
            data_type: Type of data to process ("regular", "prequal", or "both")
            
        Returns:
            List of processed file names
            
        Raises:
            ValueError: If invalid data_type provided
            RuntimeError: If processing fails
        """
        if data_type not in ["regular", "prequal", "both"]:
            raise ValueError(f"Invalid data_type: {data_type}")
        
        try:
            processed_files = []
            
            if data_type in ["regular", "both"]:
                self.logger.info("Processing regular test data...")
                regular_files = self.regular_processor.process()
                processed_files.extend(regular_files)
            
            if data_type in ["prequal", "both"]:
                self.logger.info("Processing prequal test data...")
                prequal_files = self.prequal_processor.process()
                processed_files.extend(prequal_files)
            
            self.logger.info(f"Test data preparation completed: {len(processed_files)} total files")
            return processed_files
            
        except Exception as e:
            self.logger.error(f"Test data preparation failed: {e}")
            raise RuntimeError(f"Test data preparation failed: {e}") from e