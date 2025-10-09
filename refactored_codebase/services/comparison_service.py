"""
Comparison and CSV merging service for test result analysis.

Provides functionality to compare test results between different runs
and merge CSV comparison results into Excel reports.
"""

import json
import csv
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
import logging
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

from ..core.config import ConfigurationManager


class ResultComparator:
    """
    Compares test results between different runs.
    
    Provides detailed comparison functionality for JSON responses
    and generates comparison reports.
    """
    
    def __init__(self, 
                 config_manager: ConfigurationManager,
                 logger: logging.Logger):
        """
        Initialize result comparator.
        
        Args:
            config_manager: Configuration manager instance
            logger: Logger instance
        """
        self.config_manager = config_manager
        self.logger = logger
        self.path_config = config_manager.path_config
    
    def compare_json_objects(self, obj1: Any, obj2: Any, path: str = "") -> List[str]:
        """
        Recursively compare two JSON objects and return differences.
        
        Args:
            obj1: First JSON object
            obj2: Second JSON object
            path: Current path in the object hierarchy
            
        Returns:
            List of difference descriptions
        """
        differences = []
        
        if type(obj1) != type(obj2):
            differences.append(f"{path}: Type mismatch - {type(obj1).__name__} vs {type(obj2).__name__}")
            return differences
        
        if isinstance(obj1, dict):
            all_keys = set(obj1.keys()) | set(obj2.keys())
            
            for key in all_keys:
                current_path = f"{path}.{key}" if path else key
                
                if key not in obj1:
                    differences.append(f"{current_path}: Missing in first object")
                elif key not in obj2:
                    differences.append(f"{current_path}: Missing in second object")
                else:
                    differences.extend(self.compare_json_objects(obj1[key], obj2[key], current_path))
        
        elif isinstance(obj1, list):
            if len(obj1) != len(obj2):
                differences.append(f"{path}: Array length mismatch - {len(obj1)} vs {len(obj2)}")
            
            for i in range(min(len(obj1), len(obj2))):
                current_path = f"{path}[{i}]"
                differences.extend(self.compare_json_objects(obj1[i], obj2[i], current_path))
        
        else:
            if obj1 != obj2:
                differences.append(f"{path}: Value mismatch - '{obj1}' vs '{obj2}'")
        
        return differences
    
    def load_response_files(self, folder_name: str) -> Dict[str, Any]:
        """
        Load all response files from a specific folder.
        
        Args:
            folder_name: Name of the folder containing response files
            
        Returns:
            Dictionary mapping file names to response data
        """
        folder_path = self.path_config.test_response_dir / folder_name
        
        if not folder_path.exists():
            raise FileNotFoundError(f"Response folder not found: {folder_path}")
        
        responses = {}
        response_files = list(folder_path.glob("*_response.json"))
        
        for response_file in response_files:
            try:
                with open(response_file, 'r', encoding='utf-8') as file:
                    response_data = json.load(file)
                    # Use base name without _response suffix as key
                    base_name = response_file.stem.replace('_response', '')
                    responses[base_name] = response_data
            except Exception as e:
                self.logger.error(f"Failed to load response file {response_file}: {e}")
        
        self.logger.info(f"Loaded {len(responses)} response files from {folder_name}")
        return responses
    
    def compare_responses(self, pre_responses: Dict[str, Any], 
                         post_responses: Dict[str, Any]) -> Dict[str, Any]:
        """
        Compare responses between two test runs.
        
        Args:
            pre_responses: Responses from pre-test run
            post_responses: Responses from post-test run
            
        Returns:
            Dictionary containing comparison results
        """
        comparison_results = {
            'total_common_files': 0,
            'files_with_differences': 0,
            'files_identical': 0,
            'detailed_results': []
        }
        
        # Find common files
        common_files = set(pre_responses.keys()) & set(post_responses.keys())
        comparison_results['total_common_files'] = len(common_files)
        
        for file_name in common_files:
            pre_response = pre_responses[file_name].get('response_data', {})
            post_response = post_responses[file_name].get('response_data', {})
            
            # Compare responses
            differences = self.compare_json_objects(pre_response, post_response)
            
            result = {
                'file_name': file_name,
                'has_differences': len(differences) > 0,
                'difference_count': len(differences),
                'differences': differences,
                'pre_status': pre_responses[file_name].get('success', False),
                'post_status': post_responses[file_name].get('success', False)
            }
            
            comparison_results['detailed_results'].append(result)
            
            if len(differences) > 0:
                comparison_results['files_with_differences'] += 1
            else:
                comparison_results['files_identical'] += 1
        
        self.logger.info(f"Comparison completed: {comparison_results['files_identical']} identical, "
                        f"{comparison_results['files_with_differences']} with differences")
        
        return comparison_results
    
    def save_comparison_report(self, comparison_results: Dict[str, Any], 
                              pre_folder: str, post_folder: str) -> Path:
        """
        Save comparison results to CSV file.
        
        Args:
            comparison_results: Results from comparison
            pre_folder: Pre-test folder name
            post_folder: Post-test folder name
            
        Returns:
            Path to saved comparison report
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = (self.path_config.comparison_dir / 
                      f"comparison_{pre_folder}_vs_{post_folder}_{timestamp}.csv")
        
        # Ensure output directory exists
        report_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Prepare CSV data
        csv_data = []
        for result in comparison_results['detailed_results']:
            csv_data.append({
                'File Name': result['file_name'],
                'Has Differences': 'Yes' if result['has_differences'] else 'No',
                'Difference Count': result['difference_count'],
                'Pre-Test Status': 'Success' if result['pre_status'] else 'Failed',
                'Post-Test Status': 'Success' if result['post_status'] else 'Failed',
                'Differences': '; '.join(result['differences'][:5])  # Limit to first 5 differences
            })
        
        # Write CSV file
        with open(report_file, 'w', newline='', encoding='utf-8') as file:
            if csv_data:
                writer = csv.DictWriter(file, fieldnames=csv_data[0].keys())
                writer.writeheader()
                writer.writerows(csv_data)
        
        self.logger.info(f"Comparison report saved: {report_file}")
        return report_file


class CSVMerger:
    """
    Merges CSV comparison results into Excel workbooks.
    
    Provides functionality to combine multiple CSV files into
    formatted Excel reports with styling and analysis.
    """
    
    def __init__(self, 
                 config_manager: ConfigurationManager,
                 logger: logging.Logger):
        """
        Initialize CSV merger.
        
        Args:
            config_manager: Configuration manager instance
            logger: Logger instance
        """
        self.config_manager = config_manager
        self.logger = logger
        self.path_config = config_manager.path_config
    
    def find_csv_files(self, sub_folder: str) -> List[Path]:
        """
        Find all CSV files in a subfolder.
        
        Args:
            sub_folder: Subfolder name to search
            
        Returns:
            List of CSV file paths
        """
        folder_path = self.path_config.comparison_dir / sub_folder
        
        if not folder_path.exists():
            raise FileNotFoundError(f"CSV folder not found: {folder_path}")
        
        csv_files = list(folder_path.glob("*.csv"))
        self.logger.info(f"Found {len(csv_files)} CSV files in {sub_folder}")
        
        return csv_files
    
    def merge_csv_files(self, csv_files: List[Path]) -> pd.DataFrame:
        """
        Merge multiple CSV files into a single DataFrame.
        
        Args:
            csv_files: List of CSV file paths
            
        Returns:
            Merged DataFrame
        """
        merged_data = []
        
        for csv_file in csv_files:
            try:
                df = pd.read_csv(csv_file)
                df['Source File'] = csv_file.name
                merged_data.append(df)
                self.logger.debug(f"Added {len(df)} rows from {csv_file.name}")
            except Exception as e:
                self.logger.error(f"Failed to read CSV file {csv_file}: {e}")
        
        if not merged_data:
            return pd.DataFrame()
        
        merged_df = pd.concat(merged_data, ignore_index=True)
        self.logger.info(f"Merged {len(merged_df)} total rows from {len(csv_files)} files")
        
        return merged_df
    
    def create_excel_workbook(self, merged_df: pd.DataFrame, 
                             sub_folder: str) -> Path:
        """
        Create formatted Excel workbook from merged data.
        
        Args:
            merged_df: Merged DataFrame
            sub_folder: Source subfolder name
            
        Returns:
            Path to created Excel file
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = (self.path_config.merged_output_dir / 
                     f"merged_comparison_{sub_folder}_{timestamp}.xlsx")
        
        # Ensure output directory exists
        excel_file.parent.mkdir(parents=True, exist_ok=True)
        
        # Create workbook
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Create summary sheet
        self._create_summary_sheet(workbook, merged_df)
        
        # Create detailed data sheet
        self._create_data_sheet(workbook, merged_df)
        
        # Save workbook
        workbook.save(excel_file)
        self.logger.info(f"Excel workbook created: {excel_file}")
        
        return excel_file
    
    def _create_summary_sheet(self, workbook: Workbook, df: pd.DataFrame) -> None:
        """Create summary sheet with statistics."""
        summary_sheet = workbook.create_sheet("Summary")
        
        # Calculate summary statistics
        total_files = len(df)
        files_with_differences = len(df[df['Has Differences'] == 'Yes'])
        files_identical = total_files - files_with_differences
        
        # Add summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Files", total_files],
            ["Files with Differences", files_with_differences],
            ["Identical Files", files_identical],
            ["Difference Rate", f"{(files_with_differences/total_files*100):.1f}%" if total_files > 0 else "0%"],
            ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        # Write summary data
        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Style header row
                if row_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
        
        # Auto-adjust column widths
        for column in summary_sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            summary_sheet.column_dimensions[column[0].column_letter].width = max_length + 2
    
    def _create_data_sheet(self, workbook: Workbook, df: pd.DataFrame) -> None:
        """Create detailed data sheet."""
        data_sheet = workbook.create_sheet("Detailed Results")
        
        # Add DataFrame to sheet
        for row in dataframe_to_rows(df, index=False, header=True):
            data_sheet.append(row)
        
        # Style header row
        for cell in data_sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # Style data rows
        for row_idx in range(2, len(df) + 2):
            # Highlight rows with differences
            has_differences = data_sheet.cell(row=row_idx, column=2).value == 'Yes'
            
            if has_differences:
                fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                for col_idx in range(1, len(df.columns) + 1):
                    data_sheet.cell(row=row_idx, column=col_idx).fill = fill
        
        # Auto-adjust column widths
        for column in data_sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value)
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            data_sheet.column_dimensions[column[0].column_letter].width = adjusted_width


class ComparisonService:
    """
    High-level service for test result comparison and CSV merging.
    
    Orchestrates comparison and merging operations and provides
    a unified interface for result analysis.
    """
    
    def __init__(self, 
                 config_manager: ConfigurationManager,
                 logger: logging.Logger):
        """
        Initialize comparison service.
        
        Args:
            config_manager: Configuration manager instance
            logger: Logger instance
        """
        self.config_manager = config_manager
        self.logger = logger
        
        # Initialize components
        self.comparator = ResultComparator(config_manager, logger)
        self.csv_merger = CSVMerger(config_manager, logger)
    
    def compare_results(self, pre_folder: str, post_folder: str) -> Dict[str, Any]:
        """
        Compare test results between two runs.
        
        Args:
            pre_folder: Name of the pre-test folder
            post_folder: Name of the post-test folder
            
        Returns:
            Dictionary containing comparison results
        """
        try:
            self.logger.info(f"Starting comparison: {pre_folder} vs {post_folder}")
            
            # Load response files
            pre_responses = self.comparator.load_response_files(pre_folder)
            post_responses = self.comparator.load_response_files(post_folder)
            
            # Perform comparison
            comparison_results = self.comparator.compare_responses(pre_responses, post_responses)
            
            # Save comparison report
            report_file = self.comparator.save_comparison_report(
                comparison_results, pre_folder, post_folder
            )
            
            # Add report file path to results
            comparison_results['report_file'] = str(report_file)
            
            self.logger.info("Comparison completed successfully")
            return comparison_results
            
        except Exception as e:
            self.logger.error(f"Comparison failed: {e}")
            raise RuntimeError(f"Comparison failed: {e}") from e
    
    def merge_csv_results(self, sub_folder: str) -> Dict[str, Any]:
        """
        Merge CSV comparison results into Excel files.
        
        Args:
            sub_folder: Subfolder containing CSV files to merge
            
        Returns:
            Dictionary containing merge results
        """
        try:
            self.logger.info(f"Starting CSV merge for: {sub_folder}")
            
            # Find CSV files
            csv_files = self.csv_merger.find_csv_files(sub_folder)
            
            if not csv_files:
                self.logger.warning(f"No CSV files found in {sub_folder}")
                return {
                    'total_csv_files': 0,
                    'successful_merges': 0,
                    'file_groups': 0,
                    'output_folder': str(self.config_manager.path_config.merged_output_dir)
                }
            
            # Merge CSV files
            merged_df = self.csv_merger.merge_csv_files(csv_files)
            
            # Create Excel workbook
            excel_file = self.csv_merger.create_excel_workbook(merged_df, sub_folder)
            
            results = {
                'total_csv_files': len(csv_files),
                'successful_merges': 1 if not merged_df.empty else 0,
                'file_groups': 1,
                'output_folder': str(self.config_manager.path_config.merged_output_dir),
                'excel_file': str(excel_file),
                'total_rows': len(merged_df)
            }
            
            self.logger.info("CSV merge completed successfully")
            return results
            
        except Exception as e:
            self.logger.error(f"CSV merge failed: {e}")
            raise RuntimeError(f"CSV merge failed: {e}") from e